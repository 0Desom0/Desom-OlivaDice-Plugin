# -*- encoding: utf-8 -*-
"""Lanota 曲库业务逻辑与图片渲染。"""

import datetime
import difflib
import math
import random
import re
import uuid
from pathlib import Path
from typing import Any

from . import config
from . import utils

try:
    from PIL import Image, ImageDraw, ImageFilter, ImageFont

    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False

function_module_note = 'Lanota 曲库查询、随机、今日曲、计算与图片回复。'
DEFAULT_BG_COLOR = (247, 219, 255, 255)
FONT_SIZE = 24
PADDING = 30
LINE_SPACING = 5
MAX_WIDTH = 800


category_map = {
    'main': 'main',
    '主线': 'main',
    'side': 'side',
    '支线': 'side',
    'expansion': 'expansion',
    '扩展': 'expansion',
    '扩展包': 'expansion',
    '曲包': 'expansion',
    'event': 'event',
    '活动': 'event',
    '限时活动': 'event',
    'subscription': 'subscription',
    '书房': 'subscription',
    '订阅': 'subscription',
    'inf': 'subscription',
    '无限': 'subscription',
}

category_name_map = {
    'main': '主线',
    'side': '支线',
    'expansion': '曲包',
    'event': '活动',
    'subscription': '书房',
}


def load_song_data() -> list[dict[str, Any]]:
    data = utils.read_json_file(utils.get_song_list_path(), [])
    return data if isinstance(data, list) else []


def save_song_data(song_data: list[dict[str, Any]]) -> bool:
    return utils.save_json_file(utils.get_song_list_path(), song_data)


def load_alias_data() -> dict[str, list[str]]:
    data = utils.read_json_file(utils.get_song_alias_path(), {})
    return data if isinstance(data, dict) else {}


def save_alias_data(alias_data: dict[str, list[str]]) -> bool:
    return utils.save_json_file(utils.get_song_alias_path(), alias_data)


def load_table_data() -> dict[str, Any]:
    data = utils.read_json_file(utils.get_song_table_path(), {})
    return data if isinstance(data, dict) else {}


def convert_excel_table_to_json(excel_path: str | Path, sheet_name: str = 'Sheet1') -> dict[str, dict[str, str]]:
    """按原 jiaoben/table.py 规则把 Excel 定数表转换为 song_table 数据。"""
    try:
        from openpyxl import load_workbook
    except Exception as exception_object:
        raise RuntimeError('缺少依赖 openpyxl，请先安装 openpyxl 后再导入 Excel 定数表。') from exception_object

    workbook = load_workbook(excel_path, data_only=True)
    real_sheet_name = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[real_sheet_name]
    result = {}

    for row in sheet.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        data = {}
        for index in range(2, len(row), 2):
            if index + 1 < len(row) and row[index] and row[index + 1]:
                data[str(row[index]).strip()] = str(row[index + 1]).strip()
        if data:
            result[key] = data

    try:
        workbook.close()
    except Exception:
        pass
    return result


def import_excel_table_to_song_table() -> tuple[bool, str]:
    table_file_list = utils.get_excel_table_file_list()
    table_dir = utils.get_excel_table_dir()
    if not table_file_list:
        return False, f'未找到 Excel 定数表，请将唯一的 .xlsx/.xlsm 文件放入：{table_dir}'
    if len(table_file_list) > 1:
        file_name_list = [Path(item).name for item in table_file_list]
        return False, 'excel_table 文件夹内只能保留一个文件，请清理后重试：\n' + '\n'.join(file_name_list)

    excel_path = Path(table_file_list[0])
    if excel_path.suffix.lower() not in config.excel_table_extension_list:
        return False, f'不支持的文件类型：{excel_path.name}\n请仅保留一个 .xlsx 或 .xlsm 文件。'

    table_data = convert_excel_table_to_json(excel_path)
    if not table_data:
        return False, f'未从 {excel_path.name} 解析到有效定数数据，请检查表格格式。'
    if not utils.save_json_file(utils.get_song_table_path(), table_data):
        return False, '写入 song_table.json 失败，请检查插件数据目录权限。'

    chart_count = sum(len(item) for item in table_data.values() if isinstance(item, dict))
    return True, (
        'Excel 定数表转换完成。\n'
        f'来源文件：{excel_path.name}\n'
        f'乐曲条目：{len(table_data)}\n'
        f'谱面定数：{chart_count}\n'
        f'已覆盖：{utils.get_song_table_path()}'
    )


def load_user_data(bot_hash: Any = None) -> dict[str, Any]:
    data = utils.read_json_file(utils.ensure_user_data_file(bot_hash), {})
    return data if isinstance(data, dict) else {}


def save_user_data(user_data: dict[str, Any], bot_hash: Any = None) -> bool:
    return utils.save_json_file(utils.ensure_user_data_file(bot_hash), user_data)


def get_today_seed() -> int:
    return int(datetime.date.today().strftime('%Y%m%d'))


def get_user_today_song(user_id: str, bot_hash: Any = None):
    user_data = load_user_data(bot_hash)
    today_seed = get_today_seed()
    user_key = str(user_id)
    user_info = user_data.setdefault(user_key, {})

    if user_info.get('today_date') == today_seed:
        chapter = str(user_info.get('today_chapter', '')).lower()
        for song in load_song_data():
            if str(song.get('chapter', '')).lower() == chapter:
                return song

    song_data = load_song_data()
    if not song_data:
        return None
    try:
        random.seed(today_seed + int(user_id))
    except Exception:
        random.seed(today_seed)
    today_song = random.choice(song_data)
    user_info['today_chapter'] = today_song.get('chapter', '')
    user_info['today_date'] = today_seed
    save_user_data(user_data, bot_hash)
    return today_song


def get_value(value: Any) -> str:
    if value is None:
        return '未知'
    text = str(value).strip()
    if text.lower() in ['none', 'no', 'n/a', 'unknown', '未知', '', 'no info']:
        return '未知'
    return text


def format_table_constant(value: Any) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    if not text:
        return ''
    range_match = re.match(r'^(\d+(?:\.\d+)?)(\s*[~-]\s*)(\d+(?:\.\d+)?)$', text)
    if range_match:
        left, sep, right = range_match.groups()
        return f'{format_table_constant(left)}{sep}{format_table_constant(right)}'
    try:
        return f'{float(text):.1f}'
    except ValueError:
        return text


def format_compact_chart_constant(
    official_value: Any,
    folk_value: Any = None,
    fallback_value: Any = None,
) -> str:
    """同时显示大等级与定数，例如 15.3(15.4) 或 15+.5(15.6)。"""
    source_value = official_value if official_value not in [None, ''] else fallback_value
    official_text = format_table_constant(source_value)
    level_text = str(fallback_value or '').strip()
    plus_level_match = re.fullmatch(r'(\d+)\+', level_text)
    if plus_level_match:
        constant_match = re.fullmatch(
            rf'{re.escape(plus_level_match.group(1))}(?:\.(\d+))?',
            official_text,
        )
        if constant_match:
            fraction = constant_match.group(1)
            official_text = level_text if not fraction or int(fraction) == 0 else f'{level_text}.{fraction}'
    official_text = re.sub(r'(?<=\d)\.0(?=\D|$)', '', official_text)
    if not official_text:
        official_text = '未知'
    if folk_value in [None, '']:
        return official_text
    folk_text = format_table_constant(folk_value)
    folk_text = re.sub(r'(?<=\d)\.0(?=\D|$)', '', folk_text)
    return f'{official_text}({folk_text})'


def format_song_info(song: dict[str, Any]) -> str:
    """按原 nonebot 插件格式渲染乐曲信息。"""
    table_data = load_table_data()
    chapter = get_value(song.get('chapter'))
    chapter_difficulty = table_data.get(chapter, {}) if chapter else {}
    legacy_info = song.get('Legacy', {})
    difficulty = song.get('difficulty', {}) if isinstance(song.get('difficulty'), dict) else {}
    official_constant = (
        song.get('official_constant', {})
        if isinstance(song.get('official_constant'), dict)
        else {}
    )
    notes = song.get('notes', {}) if isinstance(song.get('notes'), dict) else {}

    def format_difficulty_info(diff_type: str) -> str:
        diff_str = get_value(difficulty.get(diff_type))
        official_value = official_constant.get(diff_type)
        table_key = diff_type.capitalize()
        table_diff = chapter_difficulty.get(table_key)
        constant_text = format_compact_chart_constant(official_value, table_diff, diff_str)
        return f'{constant_text} (物量: {get_value(notes.get(diff_type))})'

    def format_legacy_difficulty(diff_key: str, max_key: str) -> str:
        if not isinstance(legacy_info, dict):
            return '无信息'
        diff_value = legacy_info.get(diff_key)
        max_value = legacy_info.get(max_key)
        legacy_constant = legacy_info.get('official_constant', {})
        constant_value = None
        if isinstance(legacy_constant, dict):
            constant_value = legacy_constant.get(diff_key.removeprefix('Diff').lower())
        if diff_value or max_value:
            constant_text = format_compact_chart_constant(
                constant_value,
                legacy_info.get('folk_constant', {}).get(
                    diff_key.removeprefix('Diff').lower(),
                )
                if isinstance(legacy_info.get('folk_constant'), dict)
                else None,
                diff_value,
            )
            return f'{constant_text} (物量: {get_value(max_value)})'
        return '无信息'

    line_list = [
        '══════════ 乐曲信息 ══════════',
        f'▪ 乐曲ID: {get_value(song.get("id"))}',
        f'▪ 官方songId: {get_value(song.get("official_songid"))}',
        f'▪ 曲名: {get_value(song.get("title"))}',
        f'▪ 分类: {category_name_map.get(song.get("category"), get_value(song.get("category")))}',
        f'▪ 章节: {chapter}',
        f'▪ 曲师: {get_value(song.get("artist"))}',
        f'▪ 歌手: {get_value(song.get("vocals"))}',
        f'▪ 曲风: {get_value(song.get("genre"))}',
        f'▪ 乐曲BPM: {get_value(song.get("bpm"))}',
        f'▪ 时长: {get_value(song.get("time"))}',
        f'▪ 更新版本: {get_value(song.get("version"))}',
        '══════════ 难度信息 ══════════',
        f'▪ 谱师: {get_value(song.get("chart_design"))}',
        f'    ┌ Whisper: {format_difficulty_info("whisper")}',
        f'    ├ Acoustic: {format_difficulty_info("acoustic")}',
        f'    ├ Ultra: {format_difficulty_info("ultra")}',
        f'    └ Master: {format_difficulty_info("master")}',
    ]
    if isinstance(legacy_info, dict) and legacy_info:
        line_list.extend(
            [
                '══════════ 旧谱信息 ══════════',
                f'▪ 官方songId: {get_value(legacy_info.get("official_songid"))}',
                f'▪ 谱师: {get_value(legacy_info.get("Chart Design"))}',
                f'    ┌ Whisper: {format_legacy_difficulty("DiffWhisper", "MaxWhisper")}',
                f'    ├ Acoustic: {format_legacy_difficulty("DiffAcoustic", "MaxAcoustic")}',
                f'    ├ Ultra: {format_legacy_difficulty("DiffUltra", "MaxUltra")}',
                f'    └ Master: {format_legacy_difficulty("DiffMaster", "MaxMaster")}',
            ]
        )
    line_list.extend(
        [
            '══════════ 其他信息 ══════════',
            '▪ 全曲列表: https://lanota.fandom.com/wiki/Songs',
            f'▪ 信息来源: {get_value(song.get("source_url"))}',
            '═════════════════════════',
        ]
    )
    return '\n'.join(line_list)


def get_songs_by_category(song_data: list[dict[str, Any]], category: str) -> list[dict[str, Any]]:
    return [song for song in song_data if song.get('category') == category]


def get_songs_by_level(song_data: list[dict[str, Any]], level: str) -> list[dict[str, Any]]:
    return [
        song
        for song in song_data
        if str(song.get('difficulty', {}).get('whisper')) == level
        or str(song.get('difficulty', {}).get('acoustic')) == level
        or str(song.get('difficulty', {}).get('ultra')) == level
        or str(song.get('difficulty', {}).get('master')) == level
    ]


def calculate_search_score(search_term: str, target_str: str) -> int:
    """计算规范化模糊匹配分数；0 最相似，1000 最不相似。"""

    def normalize(value: str) -> str:
        return re.sub(r'[^\w\u4e00-\u9fff]+', '', str(value).casefold())

    query = normalize(search_term)
    target_text = str(target_str).strip()
    target_variants = [normalize(target_text)]
    main_title = re.split(r'[（(\[]', target_text, maxsplit=1)[0]
    normalized_main_title = normalize(main_title)
    if normalized_main_title and normalized_main_title not in target_variants:
        target_variants.append(normalized_main_title)
    if not query or not any(target_variants):
        return 10001

    best_ratio = 0.0
    for target in target_variants:
        ratio = difflib.SequenceMatcher(None, query, target).ratio()
        shorter_length = min(len(query), len(target))
        containment_ratio = shorter_length / max(len(query), len(target))
        is_meaningful_containment = query in target or (target in query and containment_ratio >= 0.7)
        if shorter_length >= 3 and is_meaningful_containment:
            ratio = max(ratio, 0.75 + min(0.25, containment_ratio * 0.25))
        best_ratio = max(best_ratio, ratio)
    return round((1 - best_ratio) * 1000)


def find_song_by_search_term(
    search_term: str,
    song_data: list[dict[str, Any]],
    alias_data: dict[str, list[str]] | None = None,
    max_display: int = 10,
) -> tuple[list[dict[str, Any]], str | None, int]:
    alias_data = alias_data if isinstance(alias_data, dict) else load_alias_data()
    search_text = str(search_term).strip()
    if not search_text:
        return [], None, 0

    matched_songs = [song for song in song_data if str(song.get('chapter', '')).lower() == search_text.lower()]
    match_type = '章节号匹配' if matched_songs else None

    if not matched_songs:
        matched_songs = [
            song
            for song in song_data
            if search_text.casefold()
            in {
                str(song.get('id', '')).casefold(),
                str(song.get('official_songid', '')).casefold(),
            }
        ]
        match_type = 'ID匹配' if matched_songs else None

    if not matched_songs:
        alias_matches = []
        for song in song_data:
            aliases = alias_data.get(str(song.get('title')), [])
            if search_text.lower() in [str(alias).lower() for alias in aliases]:
                alias_matches.append(song)
        matched_songs = alias_matches
        match_type = '别名匹配' if matched_songs else None

    if not matched_songs:
        matched_songs = [song for song in song_data if str(song.get('title', '')).lower() == search_text.lower()]
        match_type = '曲名匹配' if matched_songs else None

    if not matched_songs:
        # 改进：使用打分制的模糊搜索，而不是简单的字符串包含
        scored_songs = []
        
        # 为每首歌计算标题和别名的最佳匹配分数
        for song in song_data:
            best_score = 10001
            search_source = None
            
            # 检查标题匹配
            title = str(song.get('title', '')).strip()
            if title:
                title_score = calculate_search_score(search_text, title)
                if title_score < best_score:
                    best_score = title_score
                    search_source = f'曲名({title})'
            
            # 检查别名匹配
            aliases = alias_data.get(str(song.get('title')), [])
            for alias in aliases:
                alias_score = calculate_search_score(search_text, str(alias).strip())
                if alias_score < best_score:
                    best_score = alias_score
                    search_source = f'别名({alias})'
            
            if best_score <= 450:
                scored_songs.append((best_score, song, search_source))
        
        scored_songs.sort(key=lambda x: x[0])
        if scored_songs:
            score_limit = min(450, scored_songs[0][0] + 150)
            scored_songs = [item for item in scored_songs if item[0] <= score_limit]
        matched_songs = [song for _, song, _ in scored_songs]
        match_type = '打分制模糊搜索' if matched_songs else None

    total_count = len(matched_songs)
    return matched_songs[:max_display], match_type, total_count


def find_artist_by_search_term(search_term: str, song_data: list[dict[str, Any]], max_display: int = 10):
    search_lower = str(search_term).strip().lower()
    if not search_lower:
        return [], None, 0
    
    # 构建不重复的曲师列表
    artist_map = {}
    for song in song_data:
        artist = str(song.get('artist', '')).strip()
        if artist:
            artist_map.setdefault(artist.lower(), artist)
    artists = list(artist_map.values())
    
    # 尝试精确匹配
    matched = [artist for artist in artists if artist.lower() == search_lower]
    match_type = '曲师精确匹配' if matched else None
    
    # 如果精确匹配失败，使用打分制模糊搜索
    if not matched:
        scored_artists = []
        for artist in artists:
            score = calculate_search_score(search_term, artist)
            if score <= 450:
                scored_artists.append((score, artist))
        
        scored_artists.sort(key=lambda x: x[0])
        if scored_artists:
            score_limit = min(450, scored_artists[0][0] + 150)
            scored_artists = [item for item in scored_artists if item[0] <= score_limit]
        matched = [artist for _, artist in scored_artists]
        match_type = '曲师打分制模糊匹配' if matched else None
    
    return matched[:max_display], match_type, len(matched)


def random_index(max_index: int) -> int:
    try:
        import requests

        url = (
            'https://www.random.org/integers/'
            f'?num=1&min=0&max={max_index}&col=1&base=10&format=plain&rnd=new'
        )
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            return int(response.text.strip())
    except Exception:
        pass
    return random.randint(0, max_index)


def parse_color(value: str):
    color_text = str(value).strip().lstrip('#')
    if not re.match(r'^[0-9a-fA-F]{6}$', color_text):
        return None
    return tuple(int(color_text[index : index + 2], 16) for index in (0, 2, 4)) + (255,)


def get_user_bg_color(user_id: str, bot_hash: Any = None):
    user_info = load_user_data(bot_hash).get(str(user_id), {})
    color = parse_color(user_info.get('bg_color', ''))
    return color or DEFAULT_BG_COLOR


def is_dark_color(color) -> bool:
    return (0.299 * color[0] + 0.587 * color[1] + 0.114 * color[2]) < 128


def get_font():
    if not PIL_AVAILABLE:
        return None
    font_path = utils.get_font_path()
    try:
        if Path(font_path).exists():
            return ImageFont.truetype(font_path, FONT_SIZE)
    except Exception:
        pass
    return ImageFont.load_default()


def wrap_text(text: str, max_chars: int = 20) -> list[str]:
    lines = []
    token_pattern = re.compile(
        r'(\d+[\+\-\*/=]+\d+|'
        r"[a-zA-Z_]+(?:'[a-zA-Z_]+)*|"
        r'\d+|'
        r'[^\w\s\u4e00-\u9fff]|'
        r'[\u4e00-\u9fff\u3000-\u303f\uff00-\uffef]|'
        r'\s)'
    )

    for paragraph in str(text).split('\n'):
        if not paragraph.strip():
            lines.append('\n')
            continue

        token_list = []
        last_end = 0
        for match in token_pattern.finditer(paragraph):
            if match.start() > last_end:
                token_list.extend(list(paragraph[last_end : match.start()]))
            token_list.append(match.group())
            last_end = match.end()
        if last_end < len(paragraph):
            token_list.extend(list(paragraph[last_end:]))

        current_line = []
        current_length = 0
        for token in token_list:
            token_length = get_text_display_length(token)
            if token_length > max_chars:
                if current_line:
                    lines.append(''.join(current_line))
                    current_line = []
                    current_length = 0
                token_parts = split_text_by_display_length(token, max_chars)
                lines.extend(token_parts[:-1])
                if token_parts:
                    current_line = [token_parts[-1]]
                    current_length = get_text_display_length(token_parts[-1])
                continue
            if current_length + token_length <= max_chars:
                current_line.append(token)
                current_length += token_length
            else:
                if current_line:
                    lines.append(''.join(current_line))
                current_line = [token]
                current_length = token_length
        if current_line:
            lines.append(''.join(current_line))
    return lines


def get_text_display_length(text: str) -> int:
    """按字体视觉宽度估算行长，约 2-3 个 ASCII 字符等于一个中文字符。"""
    source = str(text)
    if not source:
        return 0
    display_length = 0
    ascii_run_length = 0
    for character in source:
        if character.isascii():
            ascii_run_length += 1
            continue
        if ascii_run_length:
            display_length += max(1, round(ascii_run_length / 2.5))
            ascii_run_length = 0
        display_length += 1
    if ascii_run_length:
        display_length += max(1, round(ascii_run_length / 2.5))
    return display_length


def split_text_by_display_length(text: str, max_length: int) -> list[str]:
    source = str(text)
    parts = []
    current_part = ''
    for character in source:
        candidate = current_part + character
        if current_part and get_text_display_length(candidate) > max_length:
            parts.append(current_part)
            current_part = character
        else:
            current_part = candidate
    if current_part:
        parts.append(current_part)
    return parts


def cubic_bezier(t: float, p0: float, p1: float, p2: float, p3: float) -> float:
    u = 1 - t
    return u**3 * p0 + 3 * u**2 * t * p1 + 3 * u * t**2 * p2 + t**3 * p3


def get_page_items(items: list[Any], page_index: int, page_size: int) -> list[Any]:
    """获取指定页的项目。"""
    start_index = page_index * page_size
    end_index = start_index + page_size
    return items[start_index:end_index]


def format_search_results_with_pagination(results: list[dict[str, Any]], page_index: int, page_size: int) -> tuple[str, int, int]:
    """格式化搜索结果为可显示的文本，包含序号。
    
    返回：(格式化文本, 总页数, 当前页索引)
    """
    total_count = len(results)
    total_pages = max(1, math.ceil(total_count / page_size))
    page_index = max(0, min(page_index, total_pages - 1))
    
    page_results = get_page_items(results, page_index, page_size)
    start_index = page_index * page_size
    
    result_text = f'第 {page_index + 1}/{total_pages} 页：\n'
    
    for idx, song in enumerate(page_results):
        display_index = start_index + idx + 1
        chapter = song.get('chapter', '?')
        title = song.get('title', '未知')
        song_id = song.get('id', '?')
        
        result_text += f'{display_index}. {chapter} - {title} (ID: {song_id})\n'
    
    result_text = result_text.rstrip('\n')
    
    if total_count > page_size:
        result_text += '\n\n【输入序号查看详情 | 下一页/上一页/第X页 | 结束】'
    else:
        result_text += '\n\n【输入序号查看详情 | 结束】'
    
    return result_text, total_pages, page_index


def get_line_size(draw, line: str, font) -> tuple[int, int]:
    bbox = draw.textbbox((0, 0), line, font=font)
    return bbox[2] - bbox[0], bbox[3] - bbox[1]


def cleanup_image_cache() -> None:
    try:
        files = sorted(
            (
                item
                for item in Path(utils.get_generate_image_dir()).glob('lanota_*')
                if item.suffix.casefold() in {'.png', '.webp'}
            ),
            key=lambda item: item.stat().st_mtime,
        )
        for file_path in files[:-config.image_cache_limit]:
            file_path.unlink()
    except Exception:
        pass


def create_text_image(text: str, user_id: str = '', max_chars: int | None = None, bot_hash: Any = None) -> str | None:
    if not PIL_AVAILABLE:
        return None
    cleanup_image_cache()
    font = get_font()
    max_chars = max_chars or config.image_max_chars
    lines = wrap_text(text, max_chars=max_chars)
    dummy = Image.new('RGB', (1, 1))
    draw = ImageDraw.Draw(dummy)
    content_width = 0
    content_height = 0
    for line in lines:
        if line == '\n':
            content_height += FONT_SIZE + LINE_SPACING
            continue
        line_width, line_height = get_line_size(draw, line, font)
        content_width = max(content_width, line_width)
        content_height += line_height + LINE_SPACING
    canvas_width = content_width + PADDING * 2
    canvas_height = content_height + PADDING * 2

    bg_color = get_user_bg_color(user_id, bot_hash)
    dark_mode = is_dark_color(bg_color)
    end_color = (0, 0, 0, 255) if dark_mode else (255, 255, 255, 255)
    image = Image.new('RGBA', (canvas_width, canvas_height), end_color)
    center_x, center_y = canvas_width // 2, canvas_height // 2
    max_radius = math.sqrt(center_x**2 + center_y**2)
    for step in range(256, 0, -1):
        progress = cubic_bezier(step / 256, 0, 0.2, 0.8, 1.0)
        radius = int(max_radius * progress)
        color = tuple(int(bg_color[i] + (end_color[i] - bg_color[i]) * progress) for i in range(3)) + (255,)
        if radius > 0:
            ImageDraw.Draw(image).ellipse(
                (center_x - radius, center_y - radius, center_x + radius, center_y + radius),
                fill=color,
            )
    for _index in range(3):
        image = image.filter(ImageFilter.GaussianBlur(radius=1))

    draw = ImageDraw.Draw(image)
    text_color = (255, 255, 255, 255) if dark_mode else (0, 0, 0, 255)
    outline_color = (0, 0, 0, 255) if dark_mode else (255, 255, 255, 255)
    y = PADDING
    for line in lines:
        if line == '\n':
            y += FONT_SIZE + LINE_SPACING
            continue
        _line_width, line_height = get_line_size(draw, line, font)
        for dx in [-1, 0, 1]:
            for dy in [-1, 0, 1]:
                if dx != 0 or dy != 0:
                    draw.text((PADDING + dx, y + dy), line, font=font, fill=outline_color)
        draw.text((PADDING, y), line, font=font, fill=text_color)
        y += line_height + LINE_SPACING

    output_dir = Path(utils.get_generate_image_dir())
    output_path = output_dir / f'lanota_{uuid.uuid4().hex[:10]}.webp'
    try:
        image.convert('RGB').save(output_path, format='WEBP', quality=95, method=6)
        return str(output_path)
    except Exception:
        fallback_path = output_path.with_suffix('.png')
        image.convert('RGB').save(fallback_path, format='PNG', optimize=True)
        return str(fallback_path)


def build_update_report(result: dict[str, Any]) -> str:
    message = '乐曲数据更新完成！\n'
    message += f'原有乐曲: {result.get("before", 0)}首\n'
    message += '\n【官方 Portal 对标】\n'
    message += f'已匹配: {result.get("official_matched", 0)}首\n'
    message += f'Legacy 已匹配: {result.get("official_legacy_matched", 0)}首\n'
    message += f'本次更新官方字段: {result.get("official_updated", 0)}首\n'
    official_pending = result.get('official_pending') or []
    message += f'待人工确认: {len(official_pending)}首\n'
    if official_pending:
        for item in official_pending[:20]:
            message += f'• {item.get("chapter", "?")} {item.get("title", "?")}\n'

    missing_songs = result.get('missing_songs', 0)
    missing_updated = result.get('missing_updated', 0)
    missing_results = result.get('missing_results', [])

    message += '\n【缺失信息更新】\n'
    if missing_songs > 0:
        message += f'待更新: {missing_songs}首\n'
        message += f'成功更新: {missing_updated}首\n'
        if missing_results:
            message += '\n详细结果:\n'
            for item in missing_results:
                status = '✓' if item.get('success') else '✗'
                missing_text = ', '.join(str(field) for field in item.get('missing', [])) or '无'
                updated_text = ', '.join(str(field) for field in item.get('updated', [])) or '无'
                message += f'{status} {item.get("title", "")}\n'
                message += f'  缺失: {missing_text}\n'
                message += f'  已更新: {updated_text}\n'
    else:
        message += '✓ 所有歌曲信息完整\n'

    message += '\n【新增乐曲】\n'
    message += f'新增: {result.get("added", 0)}首\n'
    added_titles = result.get('added_titles') or []
    if added_titles:
        message += '\n新增曲目:\n' + '\n'.join(str(title) for title in added_titles[:30])
        if len(added_titles) > 30:
            message += f'\n……共{len(added_titles)}首'
    message += f'\n\n【总计】\n当前总乐曲: {result.get("total", 0)}首'
    return message


def build_full_check_report(result: dict[str, Any]) -> str:
    """生成全量检测覆盖报告。"""
    apply_mode = bool(result.get('apply', False))
    message = '曲库全量覆盖完成！\n' if apply_mode else '曲库全量检测完成！\n'
    message += f'检测歌曲: {result.get("checked", 0)}首\n'
    if apply_mode:
        message += f'已覆盖更新: {result.get("updated", 0)}首\n'
    else:
        message += f'检测到变化: {result.get("updated", 0)}首\n'
    message += f'新增歌曲: {result.get("added", 0)}首\n'
    message += f'无变化: {result.get("unchanged", 0)}首\n'
    message += f'失败: {result.get("failed", 0)}首\n'
    message += f'官方 songId 匹配: {result.get("official_matched", 0)}首\n'
    message += f'Legacy songId 匹配: {result.get("official_legacy_matched", 0)}首\n'
    message += f'官方字段变化: {result.get("official_updated", 0)}首\n'
    official_pending = result.get('official_pending') or []
    message += f'官方匹配待确认: {len(official_pending)}首\n'
    if apply_mode:
        message += '说明: Fandom 覆盖元数据，Portal 覆盖已有歌曲及 Legacy 的官方 ID、难度和官方定数；新曲不自动匹配 ID。\n'
    else:
        message += '说明: 当前仅检测 Fandom/Portal 与本地差异及新增歌曲，未写入本地。\n'

    if official_pending:
        message += '\n【官方匹配待确认】\n'
        for item in official_pending[:20]:
            message += f'• {item.get("chapter", "?")} {item.get("title", "?")}\n'

    added_titles = [str(title) for title in (result.get('added_titles') or []) if str(title).strip()]
    if added_titles:
        message += '\n【新增歌曲】\n'
        for title in added_titles[:30]:
            message += f'• {title}\n'
        if len(added_titles) > 30:
            message += f'……共{len(added_titles)}首新增\n'

    changed_items = [
        item for item in (result.get('results') or [])
        if item.get('success') and item.get('changed')
    ]
    failed_items = [item for item in (result.get('results') or []) if not item.get('success')]

    if changed_items:
        message += '\n【已覆盖更新的曲目】\n' if apply_mode else '\n【检测到变化的曲目】\n'
        for item in changed_items[:40]:
            chapter = item.get('chapter') or '?'
            title = item.get('title') or '?'
            fields = ', '.join(str(field) for field in item.get('changed', [])[:12])
            more = '' if len(item.get('changed', [])) <= 12 else '...'
            message += f'• {chapter} {title}\n  字段: {fields}{more}\n'
        if len(changed_items) > 40:
            message += f'……共{len(changed_items)}首有变化\n'

    if failed_items:
        message += '\n【失败曲目】\n'
        for item in failed_items[:20]:
            chapter = item.get('chapter') or '?'
            title = item.get('title') or '?'
            err = item.get('error') or '未知错误'
            message += f'• {chapter} {title}\n  原因: {err}\n'
        if len(failed_items) > 20:
            message += f'……共{len(failed_items)}首失败\n'

    if apply_mode:
        message += f'\n当前总乐曲: {result.get("total", 0)}首'
    else:
        message += (
            f'\n当前本地总乐曲: {result.get("total", 0)}首\n'
            f'应用后预计总乐曲: {result.get("projected_total", result.get("total", 0))}首'
        )
    return message
